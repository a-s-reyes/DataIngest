from pathlib import Path
from typing import Any

import pytest
from sqlalchemy import create_engine, text

from dataingest.config import Mapping
from dataingest.infer import dump_mapping, infer_mapping
from dataingest.pipeline import Pipeline


def _write(path: Path, text: str) -> Path:
    path.write_text(text, encoding="utf-8")
    return path


def test_infer_int_column(tmp_path: Path) -> None:
    csv = _write(tmp_path / "x.csv", "id,age\nA,42\nB,17\nC,99\n")
    m = infer_mapping(csv)
    assert m["fields"]["age"]["type"] == "int"
    assert m["fields"]["age"]["cleaners"] == ["strip", "parse_int"]


def test_infer_decimal_column(tmp_path: Path) -> None:
    csv = _write(tmp_path / "x.csv", "id,amount\nA,1.50\nB,2.75\nC,99\n")
    m = infer_mapping(csv)
    assert m["fields"]["amount"]["type"] == "decimal"
    assert m["fields"]["amount"]["cleaners"] == ["strip", "parse_decimal"]


def test_infer_iso_date_column(tmp_path: Path) -> None:
    csv = _write(tmp_path / "x.csv", "id,d\nA,2026-01-15\nB,2026-02-28\n")
    m = infer_mapping(csv)
    assert m["fields"]["d"]["type"] == "date"
    assert m["fields"]["d"]["cleaners"] == ["strip", "parse_date_iso"]


def test_infer_us_date_column(tmp_path: Path) -> None:
    csv = _write(tmp_path / "x.csv", "id,d\nA,1/15/2026\nB,12/31/2026\n")
    m = infer_mapping(csv)
    assert m["fields"]["d"]["type"] == "date"
    assert m["fields"]["d"]["cleaners"] == ["strip", "parse_date_us"]


def test_infer_iso_datetime_column(tmp_path: Path) -> None:
    csv = _write(
        tmp_path / "x.csv",
        "id,t\nA,2026-04-12T14:22:01\nB,2026-04-12T14:22:02\n",
    )
    m = infer_mapping(csv)
    assert m["fields"]["t"]["type"] == "datetime"
    assert m["fields"]["t"]["cleaners"] == ["strip", "parse_datetime_iso"]


def test_infer_bool_literal_column(tmp_path: Path) -> None:
    csv = _write(tmp_path / "x.csv", "id,active\nA,true\nB,false\nC,yes\n")
    m = infer_mapping(csv)
    assert m["fields"]["active"]["type"] == "bool"


def test_infer_falls_back_to_str_on_mixed(tmp_path: Path) -> None:
    csv = _write(tmp_path / "x.csv", "id,mixed\nA,42\nB,hello\nC,2026-01-01\n")
    m = infer_mapping(csv)
    assert m["fields"]["mixed"]["type"] == "str"


def test_infer_empty_column_defaults_to_str(tmp_path: Path) -> None:
    csv = _write(tmp_path / "x.csv", "id,blank\nA,\nB,\n")
    m = infer_mapping(csv)
    assert m["fields"]["blank"]["type"] == "str"


def test_required_when_no_empty_samples(tmp_path: Path) -> None:
    csv = _write(tmp_path / "x.csv", "id,name\nA,alpha\nB,beta\n")
    m = infer_mapping(csv)
    assert m["fields"]["name"]["required"] is True


def test_not_required_when_some_empty(tmp_path: Path) -> None:
    csv = _write(tmp_path / "x.csv", "id,name\nA,alpha\nB,\nC,gamma\n")
    m = infer_mapping(csv)
    assert m["fields"]["name"]["required"] is False


def test_primary_key_picks_first_unique_column(tmp_path: Path) -> None:
    csv = _write(
        tmp_path / "x.csv",
        "category,sku,name\nA,SKU-1,alpha\nA,SKU-2,beta\nB,SKU-3,gamma\n",
    )
    m = infer_mapping(csv)
    assert m["target"]["primary_key"] == "sku"


def test_primary_key_falls_back_to_first_column(tmp_path: Path) -> None:
    csv = _write(tmp_path / "x.csv", "a,b,c\nx,y,z\nx,y,z\nx,y,z\n")
    m = infer_mapping(csv)
    assert m["target"]["primary_key"] == "a"


def test_primary_key_skips_columns_with_empty_values(tmp_path: Path) -> None:
    csv = _write(
        tmp_path / "x.csv",
        "maybe_id,real_id\n,A\nfoo,B\nbar,C\n",
    )
    m = infer_mapping(csv)
    assert m["target"]["primary_key"] == "real_id"


def test_mapping_uses_filename_stem_as_default_name(tmp_path: Path) -> None:
    csv = _write(tmp_path / "qualification_runs.csv", "id\nA\n")
    m = infer_mapping(csv)
    assert m["name"] == "qualification_runs"
    assert m["target"]["table"] == "qualification_runs"


def test_explicit_name_and_table_override(tmp_path: Path) -> None:
    csv = _write(tmp_path / "x.csv", "id\nA\n")
    m = infer_mapping(csv, name="my-mapping", table="custom_table")
    assert m["name"] == "my-mapping"
    assert m["target"]["table"] == "custom_table"


def test_dump_mapping_yields_loadable_yaml(tmp_path: Path) -> None:
    csv = _write(
        tmp_path / "data.csv",
        "id,name,amount\nA,alpha,1.50\nB,beta,2.75\n",
    )
    m = infer_mapping(csv)
    yaml_text = dump_mapping(m)
    yml_path = tmp_path / "out.yml"
    yml_path.write_text(yaml_text, encoding="utf-8")
    loaded = Mapping.from_yaml(yml_path)
    assert loaded.name == "data"
    assert "amount" in loaded.fields


def test_dump_mapping_never_emits_yaml_anchors(tmp_path: Path) -> None:
    csv = _write(
        tmp_path / "wide.csv",
        "a,b,c,d,e,f\n1,2,3,4,5,6\nx,y,z,u,v,w\n",
    )
    yaml_text = dump_mapping(infer_mapping(csv))
    assert "&id" not in yaml_text
    assert "*id" not in yaml_text
    assert yaml_text.count("- strip") >= 6


def test_infer_then_run_succeeds_without_manual_edits(tmp_path: Path) -> None:
    csv = _write(
        tmp_path / "telemetry_lite.csv",
        "record_id,channel,value,unit,recorded_at\n"
        "TM-1,ACC_X,0.4823,g,2026-04-12T14:22:01\n"
        "TM-2,ACC_Y,-0.0117,g,2026-04-12T14:22:01\n"
        "TM-3,ACC_Z,1.0024,g,2026-04-12T14:22:01\n",
    )
    mapping_dict = infer_mapping(csv)
    yml_path = tmp_path / "telemetry_lite.yml"
    yml_path.write_text(dump_mapping(mapping_dict), encoding="utf-8")

    db_path = tmp_path / "out.db"
    pipeline = Pipeline(
        source_uri=f"csv:///{csv.as_posix()}",
        sink_uri=f"sqlite:///{db_path.as_posix()}",
        mapping=Mapping.from_yaml(yml_path),
        error_log=tmp_path / "errors.jsonl",
    )
    result = pipeline.run()
    assert result.rows_in == 3
    assert result.rows_ok == 3
    assert result.rows_failed == 0

    engine = create_engine(f"sqlite:///{db_path}")
    with engine.connect() as conn:
        count = conn.execute(text("SELECT COUNT(*) FROM telemetry_lite")).scalar()
        assert count == 3


def test_infer_raises_on_empty_file(tmp_path: Path) -> None:
    csv = _write(tmp_path / "empty.csv", "")
    with pytest.raises(ValueError, match="no header"):
        infer_mapping(csv)


def test_infer_handles_custom_delimiter(tmp_path: Path) -> None:
    csv = _write(tmp_path / "tsv.csv", "id\tname\nA\talpha\nB\tbeta\n")
    m = infer_mapping(csv, delimiter="\t")
    assert "id" in m["fields"]
    assert "name" in m["fields"]


def test_infer_strips_utf8_bom_from_first_field(tmp_path: Path) -> None:
    csv = _write(tmp_path / "bom.csv", "﻿record_id,name\nTM-1,alpha\n")
    m = infer_mapping(csv)
    assert "record_id" in m["fields"]
    assert "﻿record_id" not in m["fields"]
    assert m["target"]["primary_key"] == "record_id"


def _write_xlsx(path: Path, rows: list[list[Any]], sheet: str = "Sheet1") -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    ws = wb.create_sheet(title=sheet)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_infer_autodetects_xlsx_from_extension(tmp_path: Path) -> None:
    xlsx = _write_xlsx(
        tmp_path / "things.xlsx",
        [
            ["id", "amount"],
            ["A", 1.5],
            ["B", 2.75],
        ],
    )
    m = infer_mapping(xlsx)
    assert m["source"]["format"] == "xlsx"
    assert "encoding" not in m["source"]
    assert "delimiter" not in m["source"]
    assert m["fields"]["amount"]["type"] == "decimal"


def test_infer_explicit_format_overrides_extension(tmp_path: Path) -> None:
    weird = tmp_path / "weird.dat"
    weird.write_text("id,name\nA,alpha\nB,beta\n", encoding="utf-8")
    m = infer_mapping(weird, format="csv")
    assert m["source"]["format"] == "csv"


def test_infer_xlsx_round_trips_through_pipeline(tmp_path: Path) -> None:
    from sqlalchemy import create_engine, text

    xlsx = _write_xlsx(
        tmp_path / "readings.xlsx",
        [
            ["record_id", "channel", "value"],
            ["TM-1", "ACC_X", 0.4823],
            ["TM-2", "ACC_Y", -0.0117],
            ["TM-3", "ACC_Z", 1.0024],
        ],
    )
    mapping_dict = infer_mapping(xlsx)
    yml_path = tmp_path / "readings.yml"
    yml_path.write_text(dump_mapping(mapping_dict), encoding="utf-8")

    db_path = tmp_path / "out.db"
    pipeline = Pipeline(
        source_uri=f"xlsx:///{xlsx.as_posix()}",
        sink_uri=f"sqlite:///{db_path.as_posix()}",
        mapping=Mapping.from_yaml(yml_path),
        error_log=tmp_path / "errors.jsonl",
    )
    result = pipeline.run()
    assert result.rows_in == 3
    assert result.rows_ok == 3

    engine = create_engine(f"sqlite:///{db_path}")
    with engine.connect() as conn:
        count = conn.execute(text("SELECT COUNT(*) FROM readings")).scalar()
        assert count == 3


def test_infer_xlsx_with_named_sheet(tmp_path: Path) -> None:
    import openpyxl

    xlsx = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    other = wb.create_sheet("Bills")
    other.append(["bill_id", "amount"])
    other.append(["B-1", 1.50])
    wb.create_sheet("Empty")
    wb.save(xlsx)

    m = infer_mapping(xlsx, sheet="Bills")
    assert m["fields"]["bill_id"]["type"] == "str"
    assert m["fields"]["amount"]["type"] == "decimal"
    assert "sheet" not in m["source"]
